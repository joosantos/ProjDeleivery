import os
from fastapi.responses import FileResponse
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from models import Insurance
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from starlette.responses import FileResponse
from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.platypus import Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.enums import TA_JUSTIFY


def create_excel_with_athletes_insurances(
    insurances: list[Insurance],
    n_results: int,
    insurance_group_name: str | None = None,
    team_name: str | None = None,
):
    original_file_path = (
        "./core/templates/excel_insurances_groups.xlsx"
        if insurance_group_name is not None
        else "./core/templates/excel_all_insurances.xlsx"
    )
    workbook = load_workbook(original_file_path)
    sheet = workbook.active

    # Save additional data
    if insurance_group_name is not None or team_name is not None:
        cell_value = ""
        if insurance_group_name is not None:
            cell_value = f"Grupo de seguros {insurance_group_name}"
        else:
            cell_value = f"Equipa {team_name}"
        sheet.cell(
            row=5,
            column=2,
            value=cell_value,
        )
    sheet.cell(
        row=4,
        column=5 if insurance_group_name is not None else 6,
        value=str(n_results),
    )

    for row_idx, insurance in enumerate(insurances, start=1):
        col_number = 1
        sheet.cell(row=8 + row_idx, column=col_number, value=row_idx)
        col_number = col_number + 1
        if insurance_group_name is None:
            sheet.cell(
                row=8 + row_idx,
                column=col_number,
                value=insurance.insured_entity.athlete.private_info.federation_number,
            )
            col_number = col_number + 1
        sheet.cell(
            row=8 + row_idx,
            column=col_number,
            value=insurance.insured_entity.athlete.name,
        )
        col_number = col_number + 1
        sheet.cell(
            row=8 + row_idx,
            column=col_number,
            value=(
                insurance.insured_entity.athlete.birthday.date()
                if insurance.insured_entity.athlete.birthday is not None
                else "-"
            ),
        )
        col_number = col_number + 1
        sheet.cell(
            row=8 + row_idx,
            column=col_number,
            value=insurance.insured_entity.athlete.private_info.nif or "-",
        )
        col_number = col_number + 1
        sheet.cell(
            row=8 + row_idx,
            column=col_number,
            value=insurance.insured_entity.athlete.private_info.identification_document.number
            or "-",
        )
    # Change font
    border_style = Side(style="thick", color="000000")
    for row in range(9, len(insurances) + 9):
        max_col_number = 5 if insurance_group_name is not None else 6
        for col in range(1, max_col_number + 1):
            cell = sheet.cell(row=row, column=col)
            cell.font = Font(name="Calibri", size=11, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if col == 1:
                continue
            if row == 9:
                if col == 2:
                    cell.border = Border(top=border_style, left=border_style)
                elif col == max_col_number:
                    cell.border = Border(top=border_style, right=border_style)
                else:
                    cell.border = Border(
                        top=border_style,
                    )
            elif row == len(insurances) + 8:
                if col == 2:
                    cell.border = Border(bottom=border_style, left=border_style)
                elif col == max_col_number:
                    cell.border = Border(bottom=border_style, right=border_style)
                else:
                    cell.border = Border(
                        bottom=border_style,
                    )
            else:
                if col == 2:
                    cell.border = Border(
                        left=border_style,
                    )
                if col == max_col_number:
                    cell.border = Border(
                        right=border_style,
                    )
    # Save the copied workbook with additional data
    final_file_path = "./core/temp/group.xlsx"
    workbook.save(final_file_path)

    # Close the workbooks
    workbook.close()
    workbook.close()
    file_name = "insurances.xlsx"
    if insurance_group_name is not None:
        file_name = f"{insurance_group_name}_insurances.xlsx"
    if team_name is not None:
        file_name = f"{team_name}_insurances.xlsx"
    return FileResponse(
        final_file_path,
        media_type="application/octet-stream",
        filename=file_name,
    )


def create_excel_with_teams_insurances(
    insurances: list[Insurance],
    n_results: int,
):
    workbook = load_workbook("./core/templates/excel_insurances_teams.xlsx")
    sheet = workbook.active

    # Save additional data
    sheet.cell(
        row=5,
        column=2,
        value="Seguros das equipas",
    )

    sheet.cell(
        row=4,
        column=5,
        value=str(n_results),
    )

    for row_idx, insurance in enumerate(insurances, start=1):
        col_number = 1
        sheet.cell(row=8 + row_idx, column=col_number, value=row_idx)
        col_number = col_number + 1
        sheet.cell(
            row=8 + row_idx,
            column=col_number,
            value=insurance.insured_entity.team.federation_number,
        )
        col_number = col_number + 1
        sheet.cell(
            row=8 + row_idx,
            column=col_number,
            value=f"{insurance.insured_entity.team.name} ({insurance.insured_entity.team.abbreviation})",
        )
    # Change font
    border_style = Side(style="thick", color="000000")
    for row in range(9, len(insurances) + 9):
        max_col_number = 6
        for col in range(1, max_col_number + 1):
            cell = sheet.cell(row=row, column=col)
            cell.font = Font(name="Calibri", size=11, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Save the copied workbook with additional data
    final_file_path = "./core/temp/group.xlsx"
    workbook.save(final_file_path)

    # Close the workbooks
    workbook.close()
    workbook.close()
    return FileResponse(
        final_file_path,
        media_type="application/octet-stream",
        filename="insurances.xlsx",
    )


def create_payment_guide(title: str, table_data: list[list[int | str]]):
    # Creating PDF
    fileName = "core/temp/payment_guide.pdf"
    pdf = canvas.Canvas(fileName, pagesize=A4)

    # Draw Image
    page_width, page_height = A4
    script_dir = os.path.dirname(__file__)
    pdf.drawImage(
        os.path.join(script_dir, f"../images/fplk_banner.webp"),
        0,
        page_height - 80,
        page_width,
        80,
        mask="auto",
    )

    # Write Title
    pdf.setFont("Helvetica-Bold", 20)
    pdf.drawCentredString(
        page_width / 2,
        page_height - 120,
        title,
    )

    # Write IBAN
    # TODO ADD IBAN AS AN ENV VAR
    pdf.setFont("Helvetica", 12)
    pdf.drawString(
        50,
        page_height - 150,
        "IBAN: 0033-0000-45424408682-05",
    )

    # Write Description
    style = getSampleStyleSheet()["Normal"]
    style.setFont = "Helvetica"
    style.textColor = colors.black
    style.fontSize = 16
    style.leading = 18
    style.alignment = TA_JUSTIFY
    table_style = TableStyle(
        [
            ("BACKGROUND", (0, 0), (-1, 0), colors.gray),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 12),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
            ("BACKGROUND", (0, 1), (-1, -1), colors.whitesmoke),
            ("GRID", (0, 0), (-1, -1), 1, colors.black),
        ]
    )

    known_cols_width = max(
        [
            sum(
                (stringWidth(col + "vvv", "Helvetica-Bold", 16))
                for col in (row[:0] + row[2:])
            )
            for row in table_data
        ]
    )
    remaining_width = page_width - 100 - known_cols_width
    known_cols_height = (len(table_data) - 1) * 20
    remaining_height = page_height - 350 - known_cols_height
    table = Table(
        table_data,
        colWidths=[None, remaining_width, None, None],
        rowHeights=[None] * (len(table_data) - 2) + [remaining_height, None],
    )
    table.setStyle(table_style)

    table.wrapOn(pdf, page_width - 100, page_height - 350)
    print(known_cols_height)
    table.drawOn(pdf, 50, 190)

    # Save PDF
    pdf.save()
    return FileResponse(
        fileName, media_type="application/octet-stream", filename=fileName
    )
